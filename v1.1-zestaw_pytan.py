# wczytywanie danych z xlsx
# based on: 
# https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
# http://zetcode.com/articles/openpyxl/

# Converting Python files to .exe:
# https://medium.com/dreamcatcher-its-blog/making-an-stand-alone-executable-from-a-python-script-using-pyinstaller-d1df9170e263

# Copyright: Szymon Butkiewicz sz.butkiewicz@gmail.com

# changelog: 
# 1.1 - selecting fields from two columns of one verse (eg. A3 and B3)


import openpyxl 
import random

print ("*-=- Witaj w programie Zestaw pytań wersja 1.1! -=-*\n")
print ("=" *30)
load_file = input("Podaj nazwę pliku z którego chcesz zczytać dane (bez rozszerzenia):\n")
save_file = input("Podaj nazwę pliku w którym chcesz zapisać zestaw pytań:\n")
items_number = input("Podaj ile pytań chcesz wylosować:\n")

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = load_file + '.xlsx')
sheets = wb.worksheets[0]
row_count = sheets.max_row		# liczba wierszy
column_count = sheets.max_column		# liczba kolumn

ws = wb.active

full_list = []		# lista z numerami wszystkich wierszy z danymi
row = 0
for i in range(row_count):
	row += 1
	full_list.append(row)

question_set = random.sample(full_list, int(items_number))		# losowanie wierszy- ilość określa zmienna items_number

book = openpyxl.Workbook()		# tworzenie nowego arkusza
sheet = book.active
new_row = 0

for q in question_set:		# przypisywanie wartości wylosowanych komórek pierwszego arkusza, do kolejnych komórek arkusza nowego
	new_row += 1
	cell_a = 'A' + str(q)	
	sheet['A' + str(new_row)] = ws[cell_a].value
	cell_b = 'B' + str(q)
	sheet['B' + str(new_row)] = ws[cell_b].value
	
book.save(save_file + ".xlsx")

print ("=" *30)
print ("Gotowe! Nowy plik powinien znajdować się w folderze z plikiem źródłowym.")